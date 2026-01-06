import csv
from io import RawIOBase, StringIO
from itertools import chain, compress, count, repeat
from os import path
from time import sleep
from typing import Self, final

import openpyxl
import openpyxl.reader.excel
import pyexcel
from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder as openpyxl_DimensionHolder

from app.utils.interruptible_io import InterruptibleIOZipFile

# monkeypatch the reference openpyxl will use for ZipFile
openpyxl.reader.excel.ZipFile = InterruptibleIOZipFile


# a sentinel argument, defined as a class so we can make typing happy
@final
class DEFAULT_ARG:
    pass


class Spreadsheet:
    class TooManyColumnsError(ValueError):
        pass

    class TooManyRowsError(ValueError):
        pass

    class AllRowsHiddenError(TooManyRowsError):
        pass

    AS_CSV_LOOP_INTERRUPTIBLE_EVERY: int = 32
    ALLOWED_FILE_EXTENSIONS = ("csv", "xlsx", "xls", "ods", "xlsm", "tsv")

    FROM_FILE_ROW_LIMIT_DEFAULT_ARG: int | None = 200_000
    COLUMN_LIMIT_FROM_HEADER_DEFAULT_ARG: bool = True
    ABSOLUTE_COLUMN_LIMIT_DEFAULT_ARG: int = 1_000
    MIN_COLUMN_LIMIT_DEFAULT_ARG: int = 50

    def __init__(self, csv_data=None, rows=None, filename="", row_limit: int | None = None):
        self.filename = filename

        if csv_data and rows:
            raise TypeError("Spreadsheet must be created from either rows or CSV data")

        self._csv_data = csv_data or ""
        self._rows = rows or []
        self._row_limit = row_limit

    @property
    def as_csv_data(self) -> str:
        if not self._csv_data:
            with StringIO() as converted:
                output = csv.writer(converted)
                for i, row in enumerate(self._rows):
                    if self._row_limit is not None and i > self._row_limit:
                        raise self.TooManyRowsError(f"Exceeded row limit of {self._row_limit}")

                    if not (i + 1) % self.AS_CSV_LOOP_INTERRUPTIBLE_EVERY:
                        # all green thread libraries will monkeypatch this to yield to the event loop
                        # and the real implementation should at least drop the GIL
                        sleep(0)

                    output.writerow(row)
                self._csv_data = converted.getvalue()
        return self._csv_data

    @classmethod
    def can_handle(cls, filename):
        return cls.get_extension(filename) in cls.ALLOWED_FILE_EXTENSIONS

    @staticmethod
    def get_extension(filename):
        return path.splitext(filename)[1].lower().lstrip(".")

    @staticmethod
    def normalise_newlines(file_content):
        return "\r\n".join(file_content.read().decode("utf-8").splitlines())

    @classmethod
    def from_rows(cls, rows, filename="", row_limit: int | None = None) -> Self:
        return cls(rows=rows, filename=filename, row_limit=row_limit)

    @staticmethod
    def _openpyxl_dimension_visible(dimensions: openpyxl_DimensionHolder, index: [int, str]) -> bool:
        # test for containment before attempting access to avoid unnecessary defaultdict allocation
        return index not in dimensions or dimensions[index].hidden is False

    @classmethod
    def _from_xlsx(  # noqa C901 is bunk
        cls,
        file_content: RawIOBase,
        filename: str,
        row_limit: int | None,
        column_limit_from_header: bool,
        absolute_column_limit: int,
        min_column_limit: int,
    ) -> Self:
        """
        Extracts all non-hidden content from the first non-hidden sheet of an xlsx file in a way that is
        comparatively efficient and with reduced blocking.
        """

        book = openpyxl.load_workbook(
            file_content,
            data_only=True,
            keep_links=False,
            rich_text=False,
            # "read only" mode is cut-down in multiple ways and doesn't even parse hidden row/column
            # or merged cell information (which we want)
            read_only=False,
        )
        sheet = next((sheet for sheet in book.worksheets if sheet.sheet_state != "hidden"), None)
        if not sheet:
            return cls.from_rows(iter(()), filename)  # empty spreadsheet

        # yield to any event loop or GIL
        sleep(0)

        # the first_visible_row will be our header row (1-based index)
        try:
            first_visible_row = next(
                # shouldn't be possible to loop infinitely as long as sheet's row_dimensions are finite
                i
                for i in (range(1, row_limit + 1) if row_limit is not None else count(1))
                # test for containment before attempting access to avoid unnecessary defaultdict allocation
                if cls._openpyxl_dimension_visible(sheet.row_dimensions, i)
            )
        except StopIteration:
            raise cls.AllRowsHiddenError("Didn't find a non-hidden row before row limit reached") from None

        # yield to any event loop or GIL
        sleep(0)

        # find the rightmost nonempty, non-hidden header cell - 1-based index
        header_last_nonempty_column = 0
        # items reversed in the hopes we encounter the rightmost first, meaning cheap early conditions can
        # exclude most cells without more expensive checks
        for (row_index, col_index), cell in reversed(sheet._cells.items()):  # APIFRAGILE on sheet._cells
            if row_index == first_visible_row and header_last_nonempty_column < col_index:
                col_letter = openpyxl_get_column_letter(col_index)
                if (cls._openpyxl_dimension_visible(sheet.column_dimensions, col_letter)) and str(
                    "" if cell.value is None else cell.value
                ).strip():
                    header_last_nonempty_column = col_index

        # yield to any event loop or GIL
        sleep(0)

        column_limit = absolute_column_limit

        if column_limit_from_header:
            if header_last_nonempty_column > absolute_column_limit:
                raise cls.TooManyColumnsError(
                    f"Last non-empty header column ({header_last_nonempty_column}) "
                    f"is beyond absolute limit of {absolute_column_limit}"
                )

            column_limit = max(header_last_nonempty_column or 1, min_column_limit)

        # cheaper to access than sheet.column_dimensions, can co-iterate through it alongside row
        visible_column_map = tuple(
            cls._openpyxl_dimension_visible(sheet.column_dimensions, col_letter)
            for col_letter in (openpyxl_get_column_letter(col_index) for col_index in range(1, column_limit + 1))
        )

        # yield to any event loop or GIL
        sleep(0)

        # now we have a column_limit of some description, in another pass find the rightmost and bottom-most
        # non-empty cells that column_limit (and row_limit) wouldn't exclude so we can perhaps further reduce
        # the bounds we have to iterate through - both 1-based indexes
        max_col_within_limit = max_row_within_limit = 1
        # why items reversed? i'm banking on the dict being populated in the order the parser encountered
        # the cells and that being approximately in top-left -> bottom-right order, meaning we encounter
        # the maximums early and most iterations can skip the more expensive conditions due to the indexes
        # just being lower
        for (row_index, col_index), cell in reversed(sheet._cells.items()):  # APIFRAGILE on sheet._cells
            if (
                col_index <= column_limit
                # cheap early filter
                and (row_index > max_row_within_limit or col_index > max_col_within_limit)
                and visible_column_map[col_index - 1]
                and cls._openpyxl_dimension_visible(sheet.row_dimensions, row_index)
                and str("" if cell.value is None else cell.value).strip()
            ):
                if row_index > row_limit:
                    # fail earlier than we otherwise would to save pointless work, this calculation
                    # counting hidden rows also saves us from having to "skip" millions of them
                    raise cls.TooManyRowsError(f"Exceeded row limit of {row_limit}")

                max_row_within_limit = max(max_row_within_limit, row_index)
                max_col_within_limit = max(max_col_within_limit, col_index)

        # yield to any event loop or GIL
        sleep(0)

        # to limit the merged-cell lookup cost no matter how many ranges the document has, flatten merged-cell
        # information into a big dict keyed by coordinates (akin to sheet._cells) with values being the value
        # each cell should instead have (taken from the top left cell of the range). note 1-based coordinates
        # to match sheet._cells keys.
        max_range = openpyxl.worksheet.cell_range.CellRange(
            min_row=1, min_col=1, max_row=max_row_within_limit, max_col=max_col_within_limit
        )
        merged_cell_map = dict(
            chain.from_iterable(
                zip(range_.cells, repeat(sheet.cell(range_.min_row, range_.min_col).value), strict=False)
                for range_ in (
                    range_.intersection(max_range)
                    for range_ in sheet.merged_cells.ranges
                    if not range_.isdisjoint(max_range)
                )
            )
        )

        # yield to any event loop or GIL
        sleep(0)

        return cls.from_rows(
            (
                (
                    merged_cell_map.get((row_index, col_index), value)
                    for col_index, value in compress(enumerate(row, 1), visible_column_map)
                )
                for row_index, row in enumerate(
                    sheet.iter_rows(
                        min_row=first_visible_row,
                        max_col=max_col_within_limit,
                        max_row=max_row_within_limit,
                        values_only=True,
                    ),
                    first_visible_row,
                )
                # test for containment before attempting access to avoid unnecessary defaultdict allocation
                if row_index not in sheet.row_dimensions or sheet.row_dimensions[row_index].hidden is False
            ),
            filename,
            row_limit=row_limit,
        )

    @classmethod
    def from_file(  # noqa C901 is bunk
        cls,
        file_content: RawIOBase,
        filename: str = "",
        row_limit: int | None | type[DEFAULT_ARG] = DEFAULT_ARG,
        column_limit_from_header: bool | type[DEFAULT_ARG] = DEFAULT_ARG,
        absolute_column_limit: int | type[DEFAULT_ARG] = DEFAULT_ARG,
        min_column_limit: int | type[DEFAULT_ARG] = DEFAULT_ARG,
    ) -> Self:
        if row_limit is DEFAULT_ARG:
            row_limit = cls.FROM_FILE_ROW_LIMIT_DEFAULT_ARG
        if column_limit_from_header is DEFAULT_ARG:
            column_limit_from_header = cls.COLUMN_LIMIT_FROM_HEADER_DEFAULT_ARG
        if absolute_column_limit is DEFAULT_ARG:
            absolute_column_limit = cls.ABSOLUTE_COLUMN_LIMIT_DEFAULT_ARG
        if min_column_limit is DEFAULT_ARG:
            min_column_limit = cls.MIN_COLUMN_LIMIT_DEFAULT_ARG
        if min_column_limit > absolute_column_limit:
            raise ValueError("min_column_limit cannot be greater than absolute_column_limit")

        extension = cls.get_extension(filename)

        if extension == "csv":
            return cls(csv_data=Spreadsheet.normalise_newlines(file_content), filename=filename, row_limit=row_limit)

        if extension == "tsv":
            file_content = StringIO(Spreadsheet.normalise_newlines(file_content))

        if extension in ("xlsx", "xlsm"):
            return cls._from_xlsx(
                file_content,
                filename,
                row_limit,
                column_limit_from_header,
                absolute_column_limit,
                min_column_limit,
            )

        column_limit = -1
        if column_limit_from_header:
            original_offset = file_content.tell()
            header = next(pyexcel.iget_array(file_type=extension, file_stream=file_content, row_limit=1), ())
            file_content.seek(original_offset)

            last_nonempty_column = next((i for i, x in reversed(tuple(enumerate(header))) if str(x).strip()), None)
            if last_nonempty_column is not None:
                if last_nonempty_column >= absolute_column_limit:
                    raise cls.TooManyColumnsError(
                        f"Last non-empty header column ({last_nonempty_column}) "
                        f"is beyond absolute limit of {absolute_column_limit}"
                    )

                column_limit = max(last_nonempty_column + 1, min_column_limit)

        return cls.from_rows(
            pyexcel.iget_array(file_type=extension, file_stream=file_content, column_limit=column_limit),
            filename,
            row_limit=row_limit,
        )
